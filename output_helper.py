from openpyxl import Workbook
import csv
import os


def write_results(headers, data, file_name, out_format='csv'):
    out_dir = "output"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    if out_format == "excel":
        print("Writing to excel file....")
        wb = Workbook()
        sheet = wb.active

        # write header
        for index, h in enumerate(headers):
            cell = sheet.cell(row=1, column=index + 1)
            cell.value = h

        # write data
        for row_idx, row in enumerate(data):
            for col_idx, col in enumerate(row):
                cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)
                cell.value = col

        print("Saving file....")
        wb.save(os.path.join(out_dir, file_name + ".xlsx"))

    elif out_format == 'csv':
        with open(os.path.join(out_dir, file_name + '.csv'), 'w', encoding='UTF8', newline='') as o_file:
            csv_writer = csv.writer(o_file)
            csv_writer.writerow(headers)
            csv_writer.writerows(data)
    else:
        print("Other format")
