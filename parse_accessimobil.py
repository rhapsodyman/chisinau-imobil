import json
import math
import re

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://accesimobil.md/apartamente-vanzare/page/{}?"

apartment_type_mapping = {"new": "1", "old": "2"}  # construction_type
apartment_state_mapping = {"euro": "1", "living": "2", "white": "3"}  # condition

# parse config
with open("config.json", 'r') as j:
    config = json.loads(j.read())

filters_part = ""

for t in config["apartmentType"]:
    filters_part += "&construction_type=" + apartment_type_mapping[t]

for t in config["apartmentState"]:
    filters_part += "&condition=" + apartment_state_mapping[t]

for r in config["rooms"]:
    filters_part += "&rooms_nr=" + r

filters_part += "&price__$min={}".format(config["minPrice"])
filters_part += "&price__$max={}".format(config["maxPrice"])

# url = url + urllib.parse.quote_plus(filters_part)
url = url + filters_part
print("Url to be used " + url)

resp = requests.get(url.format('1'))

index = BeautifulSoup(resp.content, 'html.parser')

count = re.match(r'.*?(\d+) apartamente disponibile.*', resp.text, re.DOTALL | re.MULTILINE).group(1)

all_page_items = index.find_all("div", class_="rs-card")

n_pages = math.ceil(int(count) / len(all_page_items))
apartments = []

for x in range(1, n_pages + 1):
    print('Processing page # {}'.format(x))

    resp = requests.get(url.format(str(x)))
    index = BeautifulSoup(resp.content, 'html.parser')
    all_page_items = index.find_all("div", class_="rs-card")

    for item in all_page_items:

        try:
            specs = item.find('div', class_='card-specs')
            specs_list = specs.find_all('div', class_='small')

            n_rooms = specs_list[0].text
            area = specs_list[1].text.replace('m2', '').strip()
            floor = specs_list[2].text

            link = 'https://accesimobil.md' + item.find('a')['href']
            price = item.find("div", class_='price').text.replace(u'\xa0', u'')

            address = item.find("div", class_='street').text.strip()
            region, street = address.split(" ", 1)

            meter_price = float(price.replace('€', '').replace(',', '').strip()) / float(area)

            apartment_data = [region, street, price, n_rooms, area, meter_price, floor, link]
            apartments.append(apartment_data)

        except Exception as e:
            print("Some exception", e)

headers = ['Region', 'Street', 'Price', 'Rooms', 'Area', 'Meter Price', 'Floor', 'Link']

# print(headers)
# for item in apartments:
#     print(item)

print("Writing to excel file....")
wb = Workbook()
sheet = wb.active

# write header
for index, h in enumerate(headers):
    cell = sheet.cell(row=1, column=index + 1)
    cell.value = h

# write data
for row_idx, apartment_data in enumerate(apartments):
    for col_idx, param in enumerate(apartment_data):
        cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)
        cell.value = param

print("Saving file....")
wb.save("accessimobil.xlsx")
