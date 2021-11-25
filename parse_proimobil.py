import json
import math
import urllib.parse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://proimobil.md/apartamente-de-vanzare-in-chisinau?filter="

# parse config
with open("config.json", 'r') as j:
    config = json.loads(j.read())

filters_part = ""
filters_part += "category:" + ",".join(config["category"])
filters_part += ";price:{}..{}".format(config["minPrice"], config["maxPrice"])
filters_part += ";apartmentType:" + ",".join(config["apartmentType"])
filters_part += ";apartmentState:" + ",".join(config["apartmentState"])
filters_part += ";rooms:" + ",".join(config["rooms"])
filters_part += ";offerType:" + ",".join(config["offerType"])

url = url + urllib.parse.quote_plus(filters_part)
print("Url to be used " + url)

resp = requests.get(url)
index = BeautifulSoup(resp.content, 'html.parser')

count = index.find("span", class_="text-secondary").text.replace('oferte', '').strip()
all_page_items = index.find_all("div", class_=["catCard", "box-shadow"])

n_pages = math.ceil(int(count) / len(all_page_items))
apartments = []

for x in range(1, n_pages + 1):
    print('Processing page # {}'.format(x))

    resp = requests.get(url + '&page=' + str(x))
    index = BeautifulSoup(resp.content, 'html.parser')
    all_page_items = index.find_all("div", class_=["catCard", "box-shadow"])

    for item in all_page_items:
        class_val = item['class']

        if 'sold' in class_val:
            continue  # skipping sold items

        recommended = 0
        new_price = 0

        # if 'recommended' in class_val:
        #     recommended = 1
        #
        # if 'new-price' in class_val:
        #     new_price = 1

        link = 'https://proimobil.md' + item.find('a')['href']
        price = item.find("a", class_='catCard__price').text
        address = item.find("div", class_='catCard__location').text.strip()
        region, street = address.split(",", 1)

        # data = item.find_all('div', class_='catCard__data')
        data = item.select('div.catCard__data > span')
        n_rooms = data[0].text
        area = data[2].text.replace('m2', '').strip()

        meter_price = float(price.replace('€', '').replace(' ', '').replace(',', '').strip()) / float(area)

        apartment_data = [region, street, price, n_rooms, area, meter_price, link]
        apartments.append(apartment_data)

headers = ['Region', 'Street', 'Price', 'Rooms', 'Area', 'Meter Price', 'Link']

# print(headers)
# for item in apartments:
#     print(item)

print("Writing to excel file....")
wb = Workbook()
sheet = wb.active

# write header
for index, h in enumerate(headers):
    cell = sheet.cell(row=1, column=index + 1)
    cell.value = h

# write data
for row_idx, apartment_data in enumerate(apartments):
    for col_idx, param in enumerate(apartment_data):
        cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)
        cell.value = param

print("Saving file....")
wb.save("proimobil.xlsx")
